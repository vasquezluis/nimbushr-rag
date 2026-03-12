"""
Excel Loader Module
Loads Excel files (.xlsx, .xls, .csv) into structured data for RAG ingestion.

"pages" to sheets, and "sections" to column header groups.
"""

import csv
from pathlib import Path
from typing import Dict, List
from app.settings import settings

import openpyxl

# ── helpers ────────────────────────────────────────────────────────────────────


def _load_xlsx(file_path: Path) -> Dict[str, List[List]]:
    """Return {sheet_name: [[row], [row], ...]} preserving header rows."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Keep rows that have at least one non-empty cell
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                rows.append([str(cell) if cell is not None else "" for cell in row])
        if rows:
            sheets[sheet_name] = rows
    wb.close()
    return sheets


def _load_csv(file_path: Path) -> Dict[str, List[List]]:
    """Treat a CSV as a single sheet named after the file stem."""
    rows = []
    with open(file_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        for row in reader:
            if any(cell.strip() for cell in row):
                rows.append(row)
    return {file_path.stem: rows} if rows else {}


# ── public API ─────────────────────────────────────────────────────────────────


def load_excel_file(file_path: Path) -> Dict[str, List[List]]:
    """
    Load an Excel or CSV file.

    Returns:
        {sheet_name: [[row_values], ...]}
        The first row of each sheet is assumed to be the header.
    """
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _load_xlsx(file_path)
    elif suffix == ".csv":
        return _load_csv(file_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")


def load_excel_files_from_directory() -> Dict[str, Dict[str, List[List]]]:
    """
    Load all Excel / CSV files in a directory.

    Returns:
        {filename: {sheet_name: [[row], ...]}}
    """

    data_dir = settings.excel_data_dir

    extensions = ("*.xlsx", "*.xls", "*.csv")
    files = []
    for ext in extensions:
        files.extend(data_dir.glob(ext))

    if not files:
        print(f"No Excel/CSV files found in {data_dir}")
        return {}

    print(f"Found {len(files)} Excel/CSV file(s)")
    result = {}
    for fp in files:
        try:
            sheets = load_excel_file(fp)
            result[fp.name] = sheets
            total_rows = sum(len(rows) for rows in sheets.values())
            print(f"  Loaded {fp.name}: {len(sheets)} sheet(s), {total_rows} rows")
        except Exception as e:
            print(f"  Error loading {fp.name}: {e}")

    return result
